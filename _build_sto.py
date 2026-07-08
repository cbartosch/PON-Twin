"""Extract real Malang STO node/routing data into malang_sto.json and compute
the Operator-A reconciliation (OLT -> Tier 3 access STO -> Tier 2 aggregation ->
Tier 1 core). Reproducible: re-run whenever the source workbook changes."""
import openpyxl, json, os, math

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = r"C:\Users\bartosch christian\Downloads\Malang Sites (cleaned)-vShared.xlsx"


def _f(x):
    try:
        return float(str(x).strip())
    except Exception:
        return None


def haversine_km(a, b):
    R = 6371.0088
    lat1, lon1, lat2, lon2 = map(math.radians, [a[0], a[1], b[0], b[1]])
    dlat, dlon = lat2 - lat1, lon2 - lon1
    h = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return 2 * R * math.asin(math.sqrt(h))


wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

# ---- Node Data -------------------------------------------------------------
nd = wb["Node Data"]
rows = list(nd.iter_rows(min_row=2, values_only=True))
nodes = {}   # code -> node (keep the lowest tier number = highest role for dupes like ML)
for r in rows:
    code = r[0]
    if not code:
        continue
    tier = r[6]
    lat, lon = _f(r[10]), _f(r[11])
    if lat is None or lon is None:
        continue
    node = {
        "sto_code": str(code).strip(),
        "sto_name": r[4],
        "sto_name_official": r[5],
        "tier": int(tier) if tier is not None else None,
        "datel": r[3],
        "witel": r[2],
        "city": r[7],
        "address": (r[8] or "").replace("\xa0", " ").strip() if r[8] else None,
        "latitude": lat,
        "longitude": lon,
    }
    # ML appears as both tier 1 (core) and tier 2; keep the core (min tier).
    if code in nodes and nodes[code]["tier"] is not None and node["tier"] is not None:
        if node["tier"] >= nodes[code]["tier"]:
            continue
    nodes[str(code).strip()] = node

# ---- Routing Data ----------------------------------------------------------
rd = wb["Routing Data"]
routes = []
for r in rd.iter_rows(min_row=3, values_only=True):
    if r[1] is None or r[2] is None:
        continue
    routes.append({
        "from": str(r[2]).strip(), "tier_from": r[3],
        "to": str(r[6]).strip(), "tier_to": r[7],
        "route_type": r[10], "distance_km": round(_f(r[11]) or 0, 3),
        "under_10km": str(r[12]).strip().lower().startswith("less"),
    })

tier1 = [n for n in nodes.values() if n["tier"] == 1]
tier2 = [n for n in nodes.values() if n["tier"] == 2]
tier3 = [n for n in nodes.values() if n["tier"] == 3]
core = tier1[0] if tier1 else None

# ---- Tier-2 aggregation parent for every Tier-3 access node ----------------
# Prefer a routing edge to a Tier-2 node; fall back to nearest Tier-2 by geodesy.
def parent_tier2(code, latlon):
    cand = []
    for e in routes:
        if e["from"] == code and e["tier_to"] == 2:
            cand.append((e["distance_km"], e["to"]))
        if e["to"] == code and e["tier_from"] == 2:
            cand.append((e["distance_km"], e["from"]))
    if cand:
        cand.sort()
        return cand[0][1], cand[0][0], "routing"
    # geodesic fallback
    best = min(tier2, key=lambda t: haversine_km(latlon, (t["latitude"], t["longitude"])))
    return best["sto_code"], round(haversine_km(latlon, (best["latitude"], best["longitude"])), 3), "nearest"

for n in tier3:
    p, dkm, method = parent_tier2(n["sto_code"], (n["latitude"], n["longitude"]))
    n["tier2_parent"] = p
    n["tier2_parent_km"] = dkm
    n["tier2_parent_method"] = method

sto = {
    "source_file": os.path.basename(XLSX),
    "region": "REGIONAL 5 JAWA TIMUR / WITEL MALANG",
    "counts": {"tier1": len(tier1), "tier2": len(tier2), "tier3": len(tier3), "routes": len(routes)},
    "core_tier1": core,
    "tier2_aggregation": tier2,
    "tier3_access": tier3,
    "routes": routes,
}
with open(os.path.join(HERE, "malang_sto.json"), "w", encoding="utf-8") as f:
    json.dump(sto, f, indent=2, ensure_ascii=False)

# ---- Reconcile Operator A twin OLTs against real Tier-3 access STOs ---------
twin = json.load(open(os.path.join(HERE, "pon_data.json"), encoding="utf-8"))
op_a_olts = [o for o in twin["olts"] if o["operator_code"] == "A"]

recon = []
for o in op_a_olts:
    ll = (o["latitude"], o["longitude"])
    ranked = sorted(tier3, key=lambda t: haversine_km(ll, (t["latitude"], t["longitude"])))
    match = ranked[0]
    dkm = haversine_km(ll, (match["latitude"], match["longitude"]))
    recon.append({
        "twin_olt_id": o["olt_id"], "area_id": o["area_id"], "area_name": o["area_name"],
        "twin_lat": o["latitude"], "twin_lon": o["longitude"],
        "matched_sto_code": match["sto_code"], "matched_sto_name": match["sto_name_official"],
        "match_distance_km": round(dkm, 3),
        "match_confidence": "high" if dkm < 3 else ("medium" if dkm < 8 else "low"),
        "tier3_lat": match["latitude"], "tier3_lon": match["longitude"],
        "tier2_parent": match["tier2_parent"], "tier2_parent_km": match["tier2_parent_km"],
        "tier1_core": core["sto_code"] if core else None,
    })

print(f"STO nodes: T1={len(tier1)} T2={len(tier2)} T3={len(tier3)} routes={len(routes)}")
print("Tier-2 aggregation nodes:", [t["sto_code"] for t in tier2])
print("\n=== OPERATOR A RECONCILIATION ===")
for r in recon:
    print(f"{r['twin_olt_id']} ({r['area_name']})")
    print(f"   -> Tier3 {r['matched_sto_code']} {r['matched_sto_name']}  "
          f"{r['match_distance_km']} km [{r['match_confidence']}]")
    print(f"      Tier2 parent {r['tier2_parent']} ({r['tier2_parent_km']} km) -> Tier1 {r['tier1_core']}")

json.dump(recon, open(os.path.join(HERE, "_recon_preview.json"), "w"), indent=2)
print("\nWrote malang_sto.json and _recon_preview.json")
