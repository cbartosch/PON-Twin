"""Integrate the real Telkom (Operator A) active-OLT inventory (NET-02) into the
twin (pon_data.json).

The twin previously carried only 19 synthetic Operator A placeholder OLTs (STO
Tier-3 sites, no ports) and structure_note flagged "Awaiting real Operator A
input". This script replaces them with the 8,516 real OLTs from
`NET-02 - Active OLT equipment.xlsx`:

  * National update -- every OLT is inserted as operator_code 'A' with real geo,
    vendor, technology (GPON/XGSPON), BIG/MINI role and derived live/spare/total
    PON ports (aggregated per-NE across the four vendor module sheets, joined via
    the Sheet1 hostname->ne map).
  * Malang recalculation -- OLTs inside the Malang bounding box are snapped to the
    nearest of the 19 Malang STO Tier-3 area anchors; each area's operator_A port
    rollups, area totals, archetype note and the dashboard are recomputed from the
    real ports. A national per-region footprint rollup is attached as
    operator_a_source.

Run:  python _merge_net02_olts.py [--dry-run]
A backup pon_data.json.bak6 is written (unless --dry-run).
"""
import argparse
import collections
import json
import os
import shutil

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PON = os.path.join(HERE, "pon_data.json")
XLSX = r"C:\Users\bartosch christian\Downloads\NET-02 - Active OLT equipment.xlsx"

# Malang bounding box (same box used across the twin's Malang reconciliation).
BBOX = dict(lat_min=-8.6, lat_max=-7.65, lng_min=112.2, lng_max=113.05)

# Per-vendor module sheet -> (ne key col, used col, free col, total col).
VENDOR_MODULE_COLS = {
    "Modul ALU": ("ne",      "pon_port_usage",              "pon_port_idle",  "pon_port_count"),
    "Modul FH":  ("ne_name", "Number_of_the_Used_PON_Ports", "free_pon_port", "Number_of_PON_Ports"),
    "Modul HW":  ("ne",      "pon_port_usage",              "pon_port_free",  "pon_port_count"),
    "Modul ZTE": ("ne_name", "number_used_port",            "number_free_port", "number_port"),
}


def _int(v):
    try:
        return int(v or 0)
    except (TypeError, ValueError):
        return 0


def _float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_source():
    """Return (olt_rows, hostname->[live,spare,total] ports)."""
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # hostname -> ne (Sheet1: col B = hostname, col E = ne)
    s1 = wb["Sheet1"].iter_rows(values_only=True)
    next(s1)
    host2ne = {str(r[1]).strip(): str(r[4]).strip()
               for r in s1 if r[1] and r[4]}

    # ne -> [used, free, total] aggregated across vendor module sheets
    ne_ports = collections.defaultdict(lambda: [0, 0, 0])
    for sheet, (kc, uc, fc, cc) in VENDOR_MODULE_COLS.items():
        it = wb[sheet].iter_rows(values_only=True)
        cols = {c: i for i, c in enumerate(next(it))}
        for r in it:
            ne = r[cols[kc]] if kc in cols else None
            if not ne:
                continue
            p = ne_ports[str(ne).strip()]
            p[0] += _int(r[cols[uc]]) if uc in cols else 0
            p[1] += _int(r[cols[fc]]) if fc in cols else 0
            p[2] += _int(r[cols[cc]]) if cc in cols else 0

    def ports_for(hostname):
        ne = host2ne.get(hostname)
        p = ne_ports.get(ne) if ne else None
        if not p or p[2] <= 0:
            return None  # no module data -> unknown ports
        return p

    it = wb["OLT"].iter_rows(values_only=True)
    cols = {c: i for i, c in enumerate(next(it))}
    olts = []
    for r in it:
        if not r or not r[cols["hostname"]]:
            continue
        hn = str(r[cols["hostname"]]).strip()
        olts.append({
            "hostname": hn,
            "area": r[cols["area"]],
            "region": r[cols["region"]],
            "district": r[cols["district"]],
            "sto": r[cols["sto"]],
            "latitude": _float(r[cols["latitude"]]),
            "longitude": _float(r[cols["longitude"]]),
            "vendor": r[cols["vendor"]],
            "type": r[cols["type"]],
            "teknologi": r[cols["teknologi"]],
            "big_mini": r[cols["big/mini"]],
            "ports": ports_for(hn),
        })
    wb.close()
    return olts


def in_malang(o):
    la, lo = o["latitude"], o["longitude"]
    if la is None or lo is None:
        return False
    return (BBOX["lat_min"] <= la <= BBOX["lat_max"]
            and BBOX["lng_min"] <= lo <= BBOX["lng_max"])


def build_record(o, area_id, area_name, archetype, scope):
    live, spare, total = (o["ports"] or [None, None, None])
    role = "Big OLT (Access)" if str(o["big_mini"]).upper().startswith("BIG") else "Mini OLT (Access)"
    site = "Telkom BIG OLT" if role.startswith("Big") else "Telkom Mini OLT"
    return {
        "olt_id": f"A-{o['hostname']}",
        "operator": "Operator A",
        "operator_code": "A",
        "area_id": area_id,
        "area_name": area_name,
        "archetype": archetype,
        "latitude": o["latitude"],
        "longitude": o["longitude"],
        "live_pon_ports": live,
        "spare_pon_ports": spare,
        "total_pon_ports": total,
        "olt_role": role,
        "deployment_status": "Actual (Telkom NET-02)",
        "site_type": site,
        "technology": o["teknologi"],
        "power_redundancy": "Telkom (actual)",
        "hostname": o["hostname"],
        "geo_source": "NET-02 Active OLT equipment",
        "vendor": o["vendor"],
        "equipment_type": o["type"],
        "scope": scope,
        "region": o["region"],
        "district": o["district"],
        "sto": o["sto"],
        "connected_homes": None,
        "homepass": None,
        "homeconnected": None,
        "notes": "Actual Operator A OLT (Telkom NET-02 active equipment).",
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="analyse only; do not write")
    args = ap.parse_args()

    d = json.load(open(PON, encoding="utf-8"))
    src = load_source()

    mal_areas = [a for a in d["areas"] if a.get("witel") == "MALANG" and a.get("tier") == 3]
    if not mal_areas:
        raise SystemExit("No Malang tier-3 areas found in twin -- aborting.")

    def nearest_area(lat, lon):
        best, bd = None, 1e18
        for a in mal_areas:
            dd = (a["anchor_latitude"] - lat) ** 2 + (a["anchor_longitude"] - lon) ** 2
            if dd < bd:
                bd, best = dd, a
        return best

    area_by_id = {a["area_id"]: a for a in mal_areas}

    # --- 1) drop the synthetic Operator A OLTs -------------------------------
    dropped = sum(1 for o in d["olts"] if o.get("operator_code") == "A")
    d["olts"] = [o for o in d["olts"] if o.get("operator_code") != "A"]

    # --- 2) build + insert real Operator A OLTs ------------------------------
    a_live = collections.Counter()
    a_spare = collections.Counter()
    a_total = collections.Counter()
    a_count = collections.Counter()
    region_roll = collections.defaultdict(lambda: collections.Counter())
    n_malang = 0
    n_ports_known = 0
    for o in src:
        if in_malang(o):
            area = nearest_area(o["latitude"], o["longitude"])
            rec = build_record(o, area["area_id"], area["area_name"],
                               area.get("archetype", ""), "Malang")
            n_malang += 1
            if o["ports"]:
                aid = area["area_id"]
                a_live[aid] += o["ports"][0]
                a_spare[aid] += o["ports"][1]
                a_total[aid] += o["ports"][2]
            a_count[area["area_id"]] += 1
        else:
            rec = build_record(o, None, o["district"], "National (Operator A)", "National")
        if o["ports"]:
            n_ports_known += 1
        # national per-region rollup
        rr = region_roll[o["region"] or "Unspecified"]
        rr["olts"] += 1
        if o["ports"]:
            rr["live"] += o["ports"][0]
            rr["spare"] += o["ports"][1]
            rr["total"] += o["ports"][2]
        d["olts"].append(rec)

    # --- 3) recalculate Malang area Operator A rollups -----------------------
    for a in mal_areas:
        aid = a["area_id"]
        a["operator_A_olts"] = a_count[aid]
        a["operator_A_live_ports"] = a_live[aid]
        a["operator_A_spare_ports"] = a_spare[aid]
        a["operator_A_total_ports"] = a_total[aid]
        a["area_live_ports_total"] = a_live[aid] + a.get("operator_B_live_ports", 0)
        a["area_spare_ports_total"] = a_spare[aid] + a.get("operator_B_spare_ports", 0)
        base_note = a.get("notes", "").split(";")[0]
        a["notes"] = base_note + (f"; Operator A reconciled to actual Telkom NET-02 "
                                  f"({a_count[aid]} OLTs, {a_live[aid]} live ports).")

        # Dominance now measured on installed PON port CAPACITY (live ports),
        # not connected homes. Capacity = the deployed access footprint each
        # operator has in the area.
        pa = a_live[aid]
        pb = a.get("operator_B_live_ports", 0)
        cap = pa + pb
        sa = round(pa / cap, 4) if cap else 0.0
        sb = round(pb / cap, 4) if cap else 0.0
        a["operator_A_capacity_share"] = sa
        a["operator_B_capacity_share"] = sb
        a["dominance_basis"] = "port_capacity (live PON ports)"
        a["dominance_test"] = ("A >55%" if sa > 0.55 else
                               "B >55%" if sb > 0.55 else "Contested")
        a["archetype"] = ("Operator A dominant (by port capacity)" if sa > 0.55 else
                          "Operator B dominant (by port capacity)" if sb > 0.55 else
                          "Contested overlap (by port capacity)")

    # --- 4) dashboard rebuild (mirror of _merge_op_b) ------------------------
    def n(coll, code=None):
        return sum(1 for x in d[coll] if code is None or x.get("operator_code") == code)
    live_ports = sum(o.get("live_pon_ports") or 0 for o in d["olts"])
    spare_ports = sum(o.get("spare_pon_ports") or 0 for o in d["olts"])
    a_live_nat = sum(o.get("live_pon_ports") or 0 for o in d["olts"] if o.get("operator_code") == "A")
    a_spare_nat = sum(o.get("spare_pon_ports") or 0 for o in d["olts"] if o.get("operator_code") == "A")
    a_malang = sum(1 for o in d["olts"] if o.get("operator_code") == "A" and o.get("scope") == "Malang")

    # Metrics this script manages/replaces (so re-running is idempotent) plus the
    # obsolete synthetic Operator A placeholders that NET-02 supersedes.
    managed = {
        "Operator A OLTs (actual)", "Operator A OLTs in Malang",
        "Operator A live PON ports", "Operator A spare PON ports",
        "Operator A OLTs (STO Tier-3)", "Operator A synthetic data",
    }
    area_ids = {a["area_id"] for a in d["areas"]}
    head = [m for m in d["dashboard"]
            if not str(m["Metric"]).startswith("MAL-AR")
            and m["Metric"] not in area_ids
            and m["Metric"] not in managed
            and m["Metric"] != "Area"]
    new_dash = []
    for m in head:
        metric = m["Metric"]
        if metric == "OLTs":
            m["Value"] = n("olts")
            m["Comment"] = f"{n('olts','A')} A (Telkom NET-02) + {n('olts','B')} B (PLN IconPlus)"
        elif metric == "Live PON ports":
            m["Value"] = live_ports
        elif metric == "Spare PON ports":
            m["Value"] = spare_ports
        new_dash.append(m)
    new_dash += [
        {"Metric": "Operator A OLTs (actual)", "Value": n("olts", "A"), "Unit": "ea",
         "Comment": "Telkom NET-02 active OLT equipment (national)"},
        {"Metric": "Operator A OLTs in Malang", "Value": a_malang, "Unit": "ea",
         "Comment": "Snapped to 19 STO Tier-3 areas"},
        {"Metric": "Operator A live PON ports", "Value": a_live_nat, "Unit": "port",
         "Comment": "Derived from NET-02 vendor module sheets"},
        {"Metric": "Operator A spare PON ports", "Value": a_spare_nat, "Unit": "port",
         "Comment": "Derived from NET-02 vendor module sheets"},
        {"Metric": "Area", "Value": "Archetype", "Unit": "A ports (live) / B ports (live)", "Comment": ""},
    ]
    for a in mal_areas:
        new_dash.append({
            "Metric": a["area_id"],
            "Value": a.get("archetype", ""),
            "Unit": f'{a.get("operator_A_live_ports",0)} / {a.get("operator_B_live_ports",0)}',
            "Comment": f'{a.get("dominance_test", "")} '
                       f'(A {round(100*a.get("operator_A_capacity_share",0))}% cap)',
        })
    d["dashboard"] = new_dash

    # --- 5) operator A source + national footprint rollup --------------------
    by_vendor = collections.Counter(o["vendor"] for o in src)
    by_tech = collections.Counter(o["teknologi"] for o in src)
    by_size = collections.Counter(o["big_mini"] for o in src)
    d["operator_a_source"] = {
        "operator": "Telkom (Operator A)",
        "source_file": os.path.basename(XLSX),
        "counts": {
            "olts_total": len(src),
            "olts_malang": n_malang,
            "olts_with_port_data": n_ports_known,
            "by_vendor": dict(by_vendor),
            "by_technology": dict(by_tech),
            "by_size": dict(by_size),
        },
        "national_ports": {
            "live": a_live_nat, "spare": a_spare_nat,
            "total": sum(o.get("total_pon_ports") or 0 for o in d["olts"] if o.get("operator_code") == "A"),
        },
        "region_rollup": {
            reg: {"olts": c["olts"], "live_ports": c["live"],
                  "spare_ports": c["spare"], "total_ports": c["total"]}
            for reg, c in sorted(region_roll.items())
        },
        "note": ("Real Operator A = Telkom NET-02 active OLT equipment. Ports "
                 "aggregated per-NE across ALU/FH/HW/ZTE vendor module sheets via "
                 "the Sheet1 hostname->ne map; OLTs without a module match carry "
                 "null ports. Malang OLTs snapped to nearest STO Tier-3 anchor; "
                 "the rest carried as national OLT points + region rollup."),
    }

    # --- 6) structure_note refresh ------------------------------------------
    d.setdefault("structure_note", {})["operator_A"] = (
        f"Operator A = real Telkom NET-02 active OLTs ({len(src)} national, "
        f"{n_malang} in Malang snapped to STO Tier-3 areas). Replaces the 19 "
        f"synthetic STO placeholders. Ports derived from vendor module sheets.")

    # --- summary -------------------------------------------------------------
    print(f"Source OLTs: {len(src)} (Malang {n_malang}, with port data {n_ports_known})")
    print(f"Dropped synthetic A OLTs: {dropped}")
    print(f"Totals now: olts={n('olts')} (A={n('olts','A')} B={n('olts','B')})")
    print(f"Operator A national ports: live={a_live_nat} spare={a_spare_nat}")
    print("Malang area A rollups (area: OLTs A-live/spare | A/B cap share -> archetype):")
    for a in mal_areas:
        print(f"  {a['area_id']:5} {a.get('operator_A_olts',0):3} OLTs  "
              f"{a.get('operator_A_live_ports',0):5} / {a.get('operator_A_spare_ports',0):5}  | "
              f"A {a.get('operator_A_capacity_share',0):.0%} / B {a.get('operator_B_capacity_share',0):.0%}"
              f" -> {a.get('archetype','')}")

    if args.dry_run:
        print("\n[dry-run] pon_data.json NOT written.")
        return
    shutil.copyfile(PON, PON + ".bak6")
    json.dump(d, open(PON, "w", encoding="utf-8"), ensure_ascii=False)
    print(f"\nWrote {PON} (backup pon_data.json.bak6).")


if __name__ == "__main__":
    main()
